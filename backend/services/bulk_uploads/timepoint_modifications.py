"""
Bulk upload service for populating brine_modification_description on existing
experimental_results rows.

Upload file format (CSV or Excel)
----------------------------------
Required columns:
    experiment_id           – string matching experiments.experiment_id
    time_point              – numeric days (e.g. 7, 7.0, "0")
    experiment_modification – free-text description of the brine modification

Optional columns:
    timepoint_type      – "actual_day" | "bucket_day" (default: bucket_day
                          tolerant match via find_timepoint_candidates)
    overwrite_existing  – "true" / "false" (per-row; default False)

Matching logic
--------------
1.  Trim experiment_id whitespace.
2.  Look up Experiment by experiment_id; record error if not found.
3.  Convert time_point to float (handles "7", "7.0", "0", "0.0").
4.  Call find_timepoint_candidates(db, experiment.id, float(time_point))
    which applies a ±0.0001-day tolerance against both
    time_post_reaction_days and time_post_reaction_bucket_days.
5.  If timepoint_type == "actual_day", additionally require an
    exact match on time_post_reaction_days (within tolerance).
6.  Prefer is_primary_timepoint_result=1 when multiple candidates exist.
7.  If no candidate found → unmatched timepoint error.

Duplicate handling
------------------
- Pre-scan: (experiment_id, time_point) pairs that appear more than once
  in the uploaded file are flagged as duplicates.
- Default strict mode: the entire batch is rejected if any duplicates exist
  and not all of them carry overwrite_existing = true.
- If all duplicate rows carry overwrite_existing = true, last row wins and a
  warning is surfaced in the feedback.

Write logic
-----------
- Blank modification + overwrite_existing False  → skip (no change)
- Blank modification + overwrite_existing True   → clear field, set flag False
- Non-blank modification                         → set description, @validates
                                                   auto-sets has_brine_modification
- Every write is logged to ModificationsLog with old/new values and the
  source filename.

Returns
-------
(updated, skipped, errors, row_feedbacks)
  updated         – int rows updated
  skipped         – int rows skipped (no change needed / blank + no overwrite)
  errors          – List[str] global errors
  row_feedbacks   – List[dict] one dict per input row with keys:
      row, experiment_id, time_point, status, old_value, new_value,
      result_id, warnings, errors
"""
from __future__ import annotations

import io
import math
from collections import Counter
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from sqlalchemy.orm import Session

from database import Experiment, ExperimentalResults, ModificationsLog
from backend.services.result_merge_utils import (
    find_timepoint_candidates,
    normalize_timepoint,
    TIMEPOINT_TOLERANCE_DAYS,
)
from frontend.config.variable_config import (
    TIMEPOINT_MODIFICATIONS_REQUIRED_COLUMNS,
    TIMEPOINT_MODIFICATIONS_OPTIONAL_COLUMNS,
)

# ---------------------------------------------------------------------------
# Column name aliases so flexible user-supplied headers map to internals
# ---------------------------------------------------------------------------
_COLUMN_ALIASES: Dict[str, str] = {
    # experiment_id
    "experiment_id":        "experiment_id",
    "experimentid":         "experiment_id",
    "experiment id":        "experiment_id",
    # time_point
    "time_point":           "time_point",
    "time point":           "time_point",
    "time_point_(days)":    "time_point",
    "time point (days)":    "time_point",
    "time_post_reaction":   "time_point",
    "time_post_reaction_days": "time_point",
    "time (days)":          "time_point",
    "time(days)":           "time_point",
    "time days":            "time_point",
    # modification description
    "experiment_modification":   "experiment_modification",
    "experiment modification":   "experiment_modification",
    "brine_modification_description": "experiment_modification",
    "brine modification":        "experiment_modification",
    "modification":              "experiment_modification",
    "description":               "experiment_modification",
    # optional
    "timepoint_type":   "timepoint_type",
    "timepoint type":   "timepoint_type",
    "overwrite_existing": "overwrite_existing",
    "overwrite existing": "overwrite_existing",
    "overwrite":          "overwrite_existing",
}


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _parse_bool_flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not math.isnan(float(value)):
        return bool(int(value))
    if isinstance(value, str):
        return value.strip().lower() in ("true", "1", "yes")
    return False


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Remap uploaded column headers to internal canonical names."""
    # Strip asterisks from template headers
    df.columns = [str(c).replace("*", "").strip() for c in df.columns]

    ci_aliases = {k.lower(): v for k, v in _COLUMN_ALIASES.items()}
    new_cols = []
    for col in df.columns:
        mapped = _COLUMN_ALIASES.get(col) or ci_aliases.get(col.lower())
        new_cols.append(mapped if mapped else col)
    df.columns = new_cols
    return df


class TimepointModificationsUploadService:
    """Service for bulk-populating brine_modification_description on timepoint rows."""

    # ------------------------------------------------------------------
    # Public entry points
    # ------------------------------------------------------------------

    @staticmethod
    def bulk_upsert_from_file(
        db: Session,
        file_bytes: bytes,
        filename: str = "upload",
        overwrite_all: bool = False,
        dry_run: bool = False,
        modified_by: str = "system",
    ) -> Tuple[int, int, List[str], List[Dict[str, Any]]]:
        """
        Parse a CSV or Excel file and update brine_modification_description for
        each matched experimental_results row.

        Parameters
        ----------
        db            Database session.
        file_bytes    Raw bytes of the uploaded file.
        filename      Original filename (used in audit log and feedback).
        overwrite_all Global overwrite toggle; per-row overwrite_existing overrides this.
        dry_run       When True, validate and preview but do NOT persist.
        modified_by   Username for the audit log.

        Returns
        -------
        (updated, skipped, errors, row_feedbacks)
        """
        errors: List[str] = []

        # ------------------------------------------------------------------
        # 1. Parse file
        # ------------------------------------------------------------------
        try:
            df = TimepointModificationsUploadService._read_file(file_bytes, filename)
        except Exception as exc:
            return 0, 0, [f"Failed to read file: {exc}"], []

        df = _normalize_columns(df)

        # ------------------------------------------------------------------
        # 2. Validate required columns present
        # ------------------------------------------------------------------
        missing = [
            c for c in TIMEPOINT_MODIFICATIONS_REQUIRED_COLUMNS
            if c not in df.columns
        ]
        if missing:
            return 0, 0, [
                f"Missing required column(s): {', '.join(missing)}. "
                f"Expected: {', '.join(TIMEPOINT_MODIFICATIONS_REQUIRED_COLUMNS)}"
            ], []

        records = df.to_dict("records")

        # ------------------------------------------------------------------
        # 3. Parse and clean each row
        # ------------------------------------------------------------------
        cleaned: List[Dict[str, Any]] = []
        parse_feedbacks: List[Dict[str, Any]] = []

        for row_index, rec in enumerate(records):
            row_num = row_index + 2  # 1-based + header row
            fb: Dict[str, Any] = {
                "row": row_num,
                "experiment_id": "",
                "time_point": None,
                "status": "error",
                "old_value": None,
                "new_value": None,
                "result_id": None,
                "warnings": [],
                "errors": [],
            }

            # Skip entirely blank rows
            if all(_is_blank(v) for v in rec.values()):
                continue

            # --- experiment_id ---
            raw_exp_id = rec.get("experiment_id")
            if _is_blank(raw_exp_id):
                err = "Missing experiment_id."
                errors.append(f"Row {row_num}: {err}")
                fb["errors"].append(err)
                parse_feedbacks.append(fb)
                continue
            exp_id = str(raw_exp_id).strip()
            fb["experiment_id"] = exp_id

            # --- time_point ---
            raw_time = rec.get("time_point")
            if _is_blank(raw_time):
                err = "'time_point' is required (use 0 for pre-reaction)."
                errors.append(f"Row {row_num}: {err}")
                fb["errors"].append(err)
                parse_feedbacks.append(fb)
                continue
            try:
                time_val = float(raw_time)
            except (ValueError, TypeError):
                err = f"'time_point' must be a number, got '{raw_time}'."
                errors.append(f"Row {row_num}: {err}")
                fb["errors"].append(err)
                parse_feedbacks.append(fb)
                continue
            fb["time_point"] = time_val

            # --- modification text ---
            raw_mod = rec.get("experiment_modification")
            modification = None if _is_blank(raw_mod) else str(raw_mod).strip()

            # --- optional flags ---
            timepoint_type = None
            raw_tt = rec.get("timepoint_type")
            if not _is_blank(raw_tt):
                timepoint_type = str(raw_tt).strip().lower()

            raw_ow = rec.get("overwrite_existing")
            overwrite = overwrite_all if _is_blank(raw_ow) else _parse_bool_flag(raw_ow)

            fb["status"] = "parsed"
            cleaned.append({
                "_row_num": row_num,
                "experiment_id": exp_id,
                "time_point": time_val,
                "modification": modification,
                "timepoint_type": timepoint_type,
                "overwrite": overwrite,
            })

        # Merge parse error feedbacks early
        all_feedbacks: List[Dict[str, Any]] = list(parse_feedbacks)

        # ------------------------------------------------------------------
        # 4. Duplicate pre-scan
        # ------------------------------------------------------------------
        key_counts: Counter = Counter(
            (r["experiment_id"], r["time_point"]) for r in cleaned
        )
        duplicates = {k for k, cnt in key_counts.items() if cnt > 1}

        if duplicates:
            # Check if all duplicate rows have overwrite=True
            dup_rows_without_overwrite = [
                r for r in cleaned
                if (r["experiment_id"], r["time_point"]) in duplicates
                and not r["overwrite"]
            ]
            if dup_rows_without_overwrite:
                dup_summary = "; ".join(
                    f"({e}, {t})"
                    for e, t in sorted(duplicates)[:5]
                )
                errors.append(
                    f"Duplicate (experiment_id, time_point) pairs found in the upload: "
                    f"{dup_summary}. "
                    "Set overwrite_existing=true on all duplicate rows to allow "
                    "'last row wins' behaviour, or de-duplicate the file first."
                )
                # Surface duplicates in per-row feedback too
                for r in cleaned:
                    if (r["experiment_id"], r["time_point"]) in duplicates and not r["overwrite"]:
                        all_feedbacks.append({
                            "row": r["_row_num"],
                            "experiment_id": r["experiment_id"],
                            "time_point": r["time_point"],
                            "status": "error",
                            "old_value": None,
                            "new_value": None,
                            "result_id": None,
                            "warnings": [],
                            "errors": [
                                "Duplicate row rejected (set overwrite_existing=true to allow)."
                            ],
                        })
                return 0, 0, errors, all_feedbacks

        # ------------------------------------------------------------------
        # 5. Dry-run: resolve matches without writing
        # ------------------------------------------------------------------
        if dry_run:
            return TimepointModificationsUploadService._dry_run(
                db, cleaned, duplicates, all_feedbacks,
            )

        # ------------------------------------------------------------------
        # 6. Persist
        # ------------------------------------------------------------------
        return TimepointModificationsUploadService._persist(
            db, cleaned, duplicates, all_feedbacks,
            filename=filename, modified_by=modified_by,
            global_errors=errors,
        )

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _read_file(file_bytes: bytes, filename: str) -> pd.DataFrame:
        """Parse CSV or Excel bytes into a DataFrame."""
        lower = filename.lower()
        if lower.endswith(".csv"):
            return pd.read_csv(io.BytesIO(file_bytes))
        # Excel: pick first non-instruction sheet
        xls = pd.ExcelFile(io.BytesIO(file_bytes))
        target = None
        for sheet in xls.sheet_names:
            if "INSTRUCTION" not in sheet.upper():
                target = sheet
                break
        if target is None and xls.sheet_names:
            target = xls.sheet_names[0]
        return pd.read_excel(xls, sheet_name=target)

    @staticmethod
    def _resolve_result(
        db: Session,
        experiment_id: str,
        time_point: float,
        timepoint_type: Optional[str],
    ) -> Tuple[Optional[ExperimentalResults], Optional[str]]:
        """
        Look up the best ExperimentalResults candidate for (experiment_id, time_point).

        Returns (result_row, error_message).  One of them will be None.
        """
        exp = (
            db.query(Experiment)
            .filter(Experiment.experiment_id == experiment_id)
            .first()
        )
        if exp is None:
            return None, f"Experiment '{experiment_id}' not found."

        candidates = find_timepoint_candidates(db, exp.id, time_point)

        if not candidates:
            return None, (
                f"No timepoint row found for experiment '{experiment_id}' "
                f"at time_point={time_point} (±{TIMEPOINT_TOLERANCE_DAYS} day tolerance)."
            )

        # Optional: further narrow to actual_day match
        if timepoint_type == "actual_day":
            norm = normalize_timepoint(time_point)
            actual_candidates = [
                c for c in candidates
                if c.time_post_reaction_days is not None
                and abs(
                    normalize_timepoint(c.time_post_reaction_days) - norm
                ) <= TIMEPOINT_TOLERANCE_DAYS
            ]
            if actual_candidates:
                candidates = actual_candidates

        # Prefer primary timepoint result; otherwise take oldest ID
        primary = [c for c in candidates if c.is_primary_timepoint_result]
        chosen = primary[0] if primary else candidates[0]

        warning = None
        if len(candidates) > 1:
            warning = (
                f"{len(candidates)} candidate rows found for '{experiment_id}' "
                f"time_point={time_point}; using result_id={chosen.id} "
                f"(is_primary={chosen.is_primary_timepoint_result})."
            )

        return chosen, warning  # warning is not an error; caller checks type

    @staticmethod
    def _dry_run(
        db: Session,
        cleaned: List[Dict[str, Any]],
        duplicates,
        existing_feedbacks: List[Dict[str, Any]],
    ) -> Tuple[int, int, List[str], List[Dict[str, Any]]]:
        """Validate and preview without any DB writes."""
        feedbacks = list(existing_feedbacks)
        errors: List[str] = []

        for rec in cleaned:
            row_num = rec["_row_num"]
            exp_id = rec["experiment_id"]
            time_point = rec["time_point"]
            modification = rec["modification"]
            timepoint_type = rec["timepoint_type"]
            overwrite = rec["overwrite"]

            warnings: List[str] = []

            result, msg = TimepointModificationsUploadService._resolve_result(
                db, exp_id, time_point, timepoint_type,
            )

            if result is None:
                # msg is an error string here
                errors.append(f"Row {row_num}: {msg}")
                feedbacks.append({
                    "row": row_num,
                    "experiment_id": exp_id,
                    "time_point": time_point,
                    "status": "error",
                    "old_value": None,
                    "new_value": modification,
                    "result_id": None,
                    "warnings": warnings,
                    "errors": [msg],
                })
                continue

            if isinstance(msg, str):
                warnings.append(msg)

            old_val = result.brine_modification_description
            is_duplicate = (exp_id, time_point) in duplicates

            if modification is None and not overwrite:
                status = "skip"
            elif old_val and not overwrite:
                status = "would_overwrite"
            else:
                status = "dry_run_match"

            if is_duplicate:
                warnings.append("Duplicate row in upload — last-row-wins will apply.")

            feedbacks.append({
                "row": row_num,
                "experiment_id": exp_id,
                "time_point": time_point,
                "status": status,
                "old_value": old_val,
                "new_value": modification,
                "result_id": result.id,
                "warnings": warnings,
                "errors": [],
            })

        return 0, 0, errors, feedbacks

    @staticmethod
    def _persist(
        db: Session,
        cleaned: List[Dict[str, Any]],
        duplicates,
        existing_feedbacks: List[Dict[str, Any]],
        filename: str,
        modified_by: str,
        global_errors: List[str],
    ) -> Tuple[int, int, List[str], List[Dict[str, Any]]]:
        """Write matching rows and log changes to ModificationsLog."""
        feedbacks = list(existing_feedbacks)
        errors = list(global_errors)
        updated = 0
        skipped = 0

        # Keep track of last-row-wins for duplicates: process all, commit once
        for rec in cleaned:
            row_num = rec["_row_num"]
            exp_id = rec["experiment_id"]
            time_point = rec["time_point"]
            modification = rec["modification"]
            timepoint_type = rec["timepoint_type"]
            overwrite = rec["overwrite"]

            warnings: List[str] = []

            try:
                result, msg = TimepointModificationsUploadService._resolve_result(
                    db, exp_id, time_point, timepoint_type,
                )

                if result is None:
                    errors.append(f"Row {row_num}: {msg}")
                    feedbacks.append({
                        "row": row_num,
                        "experiment_id": exp_id,
                        "time_point": time_point,
                        "status": "error",
                        "old_value": None,
                        "new_value": modification,
                        "result_id": None,
                        "warnings": warnings,
                        "errors": [msg],
                    })
                    skipped += 1
                    continue

                if isinstance(msg, str):
                    warnings.append(msg)

                old_val = result.brine_modification_description

                # --- Skip / clear / write decision ---
                if modification is None:
                    if not overwrite:
                        skipped += 1
                        feedbacks.append({
                            "row": row_num,
                            "experiment_id": exp_id,
                            "time_point": time_point,
                            "status": "skipped",
                            "old_value": old_val,
                            "new_value": None,
                            "result_id": result.id,
                            "warnings": warnings,
                            "errors": [],
                        })
                        continue
                    # overwrite=True + blank → clear
                    new_val = None
                else:
                    new_val = modification

                # Protect existing value unless overwrite is set
                if old_val and old_val.strip() and not overwrite and new_val is not None:
                    skipped += 1
                    feedbacks.append({
                        "row": row_num,
                        "experiment_id": exp_id,
                        "time_point": time_point,
                        "status": "skipped",
                        "old_value": old_val,
                        "new_value": new_val,
                        "result_id": result.id,
                        "warnings": ["Existing value preserved; pass overwrite_existing=true to replace."],
                        "errors": [],
                    })
                    continue

                # Write
                result.brine_modification_description = new_val  # @validates syncs flag

                # Audit log
                try:
                    log_entry = ModificationsLog(
                        experiment_id=exp_id,
                        experiment_fk=result.experiment_fk,
                        modified_by=modified_by,
                        modification_type="update",
                        modified_table="experimental_results",
                        old_values={
                            "brine_modification_description": old_val,
                            "has_brine_modification": False if old_val is None else bool(old_val and old_val.strip()),
                        },
                        new_values={
                            "brine_modification_description": new_val,
                            "has_brine_modification": bool(new_val and str(new_val).strip()),
                            "source_filename": filename,
                            "result_id": result.id,
                        },
                    )
                    db.add(log_entry)
                except Exception:
                    pass  # Audit failure must not block the write

                is_duplicate = (exp_id, time_point) in duplicates
                if is_duplicate:
                    warnings.append("Duplicate row — last-row-wins applied.")

                updated += 1
                feedbacks.append({
                    "row": row_num,
                    "experiment_id": exp_id,
                    "time_point": time_point,
                    "status": "updated",
                    "old_value": old_val,
                    "new_value": new_val,
                    "result_id": result.id,
                    "warnings": warnings,
                    "errors": [],
                })

            except Exception as exc:
                db.rollback()
                err_msg = f"Unexpected error: {exc}"
                errors.append(f"Row {row_num}: {err_msg}")
                feedbacks.append({
                    "row": row_num,
                    "experiment_id": exp_id,
                    "time_point": time_point,
                    "status": "error",
                    "old_value": None,
                    "new_value": None,
                    "result_id": None,
                    "warnings": [],
                    "errors": [err_msg],
                })
                skipped += 1
                continue

        db.commit()
        return updated, skipped, errors, feedbacks

    # ------------------------------------------------------------------
    # Template generator
    # ------------------------------------------------------------------

    @staticmethod
    def generate_template_bytes() -> bytes:
        """Return an Excel template (.xlsx) with required and optional headers."""
        from frontend.config.variable_config import TIMEPOINT_MODIFICATIONS_TEMPLATE_HEADERS
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timepoint Modifications"

        headers = list(TIMEPOINT_MODIFICATIONS_TEMPLATE_HEADERS.values())
        required_count = len(
            [h for h in headers if h.endswith("*")]
        )

        header_fill = PatternFill("solid", fgColor="4472C4")
        optional_fill = PatternFill("solid", fgColor="70AD47")
        header_font = Font(color="FFFFFF", bold=True)

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.fill = header_fill if header.endswith("*") else optional_fill
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = max(20, len(header) + 4)

        # Instructions sheet
        inst_ws = wb.create_sheet("INSTRUCTIONS")
        instructions = [
            ("Column", "Required", "Description"),
            ("Experiment ID *", "Yes", "Must match an existing experiment_id exactly (case-sensitive)."),
            ("Time Point (days) *", "Yes", "Numeric days value (e.g. 0, 7, 14.5). Integer and decimal forms accepted."),
            ("Experiment Modification *", "Yes", "Free-text description of the brine modification at this timepoint. Leave blank to skip (or clear if overwrite=true)."),
            ("Timepoint Type", "No", "'actual_day' to match only on exact measured days. Default (blank) uses tolerant bucket matching (±0.0001 days)."),
            ("Overwrite Existing", "No", "'true' to replace existing modification text. Default is 'false' (preserve existing)."),
        ]
        for row in instructions:
            inst_ws.append(row)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
