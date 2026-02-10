import streamlit as st
import pandas as pd
import io
from database import SessionLocal, Experiment, ExperimentalResults, ScalarResults, SampleInfo, PXRFReading, ExperimentalConditions, Compound, ChemicalAdditive, AmountUnit
from database.models.enums import ExperimentStatus
from frontend.config.variable_config import SCALAR_RESULTS_TEMPLATE_HEADERS
from backend.services.bulk_uploads.chemical_inventory import ChemicalInventoryService
from backend.services.bulk_uploads.experiment_additives import ExperimentAdditivesService
from backend.services.bulk_uploads.actlabs_xrd_report import XRDUploadService
from backend.services.bulk_uploads.scalar_results import ScalarResultsUploadService
from backend.services.bulk_uploads.actlabs_titration_data import (
    ElementalCompositionService,
    ActlabsRockTitrationService,
)
from backend.services.bulk_uploads.pxrf_data import PXRFUploadService
from backend.services.bulk_uploads.rock_inventory import RockInventoryService
from backend.services.bulk_uploads.experiment_status import ExperimentStatusService
from backend.services.scalar_results_service import ScalarResultsService
from backend.services.icp_service import ICPService
from sqlalchemy.exc import IntegrityError

EXPERIMENTAL_RESULTS_REQUIRED_COLS = {
    "experiment_id", "description"  # time_post_reaction is now optional
}

def render_bulk_uploads_page():
    """
    Renders the bulk uploads page in the Streamlit app.
    Allows users to select a data type and upload data in bulk.
    """
    st.title("Bulk Uploads")

    upload_option = st.selectbox(
        "Select data type to upload:",
        (
            "Select data type",
            "New Experiments",
            "Solution Chemistry (NMR / Hydrogen / pH / Conductivity / Alkalinity)",
            "ICP-OES",
            # For EDS Later "Elemental Composition (Titration/ICP)",
            "ActLabs XRD",
            "ActLabs Rock Titration",
            "Rock Inventory",
            "pXRF Readings",
            "Chemical Inventory (Compounds)",
            "Experiment Additives (Compounds per Experiment)",
            "Update Experiment Status (ONGOING/COMPLETED)"
        )
    )

    if upload_option == "New Experiments":
        handle_new_experiments_upload()
    elif upload_option == "Solution Chemistry (NMR / Hydrogen / pH / Conductivity / Alkalinity)":
        handle_solution_chemistry_upload()
    elif upload_option == "ICP-OES":
        handle_icp_upload()
    elif upload_option == "ActLabs XRD":
        handle_xrd_upload()
    # elif upload_option == "Elemental Composition (Titration/ICP)":
    #     handle_elemental_composition_upload()
    elif upload_option == "ActLabs Rock Titration":
        handle_actlabs_titration_upload()
    elif upload_option == "Rock Inventory":
        handle_rock_samples_upload()
    elif upload_option == "pXRF Readings":
        handle_pxrf_upload()
    elif upload_option == "Chemical Inventory (Compounds)":
        upload_chemical_inventory()
    elif upload_option == "Experiment Additives (Compounds per Experiment)":
        experiments_additives()
    elif upload_option == "Update Experiment Status (ONGOING/COMPLETED)":
        handle_experiment_status_update()


def handle_xrd_upload():
    """
    Bulk upload of XRD mineralogy per sample.
    Excel format: first column is sample_id; remaining columns are mineral names; cells are amounts (typically %).
    Creates/updates ExternalAnalysis (analysis_type="XRD"), XRDAnalysis JSON, and normalized XRDPhase rows.
    """
    st.header("Bulk Upload XRD Mineralogy")
    st.markdown(
        """
        Provide an Excel with one sheet named `xrd` where:
        - The first column is `sample_id` (required)
        - Each additional column is a mineral name with percentage values per row
        The importer creates or updates XRD records per sample.
        """
    )

    # Provide template
    template_df = pd.DataFrame([
        {"sample_id": "Rock_1", "Quartz": 45.0, "Feldspar": 25.0, "Calcite": 10.0}
    ], columns=["sample_id", "Quartz", "Feldspar", "Calcite"])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        template_df.to_excel(writer, index=False, sheet_name='xrd')
        
        # Autosize columns for readability
        try:
            from frontend.components.utils import autosize_excel_columns
            autosize_excel_columns(writer, 'xrd')
        except Exception:
            pass
    buf.seek(0)
    st.download_button(
        label="Download XRD Template",
        data=buf,
        file_name="xrd_mineralogy_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    uploaded = st.file_uploader("Upload filled XRD template (xlsx)", type=["xlsx"])

    if not uploaded:
        return

    db = SessionLocal()
    try:
        created_ext, updated_ext, created_json, updated_json, created_phase, updated_phase, skipped, errors = XRDUploadService.bulk_upsert_from_excel(db, uploaded.read())
        if errors:
            db.rollback()
            st.error("Upload failed; no changes were applied.")
            for msg in errors[:50]:
                st.error(msg)
            if len(errors) > 50:
                st.info(f"...and {len(errors)-50} more errors")
            return
        db.commit()
        st.success(
            f"ExternalAnalysis (XRD) created: {created_ext}, updated: {updated_ext}. "
            f"XRDAnalysis JSON created: {created_json}, updated: {updated_json}. "
            f"XRDPhase created: {created_phase}, updated: {updated_phase}, skipped rows: {skipped}."
        )
    except Exception as e:
        db.rollback()
        st.error(f"Unexpected error during XRD upload: {e}")
    finally:
        db.close()


def handle_new_experiments_upload():
    """
    Bulk upload to create/update experiments, conditions, additives via multi-sheet Excel.
    Includes optional compounds sheet for pre-upserting inventory.
    """
    from backend.services.bulk_uploads.new_experiments import NewExperimentsUploadService
    st.header("Bulk Upload: New Experiments")

    st.markdown(
        """
        This bulk upload uses a multi-sheet Excel template with three sheets:

        1. **experiments**: experiment details like experiment ID, sample ID, date, status, and description.
        2. **conditions**: experiment conditions, such as pH, rock mass, liquid volume, temperature, etc.
        3. **additives**: chemical additives for each experiment, including compound name, amount, and unit.

        **Template notes:**
        - Columns marked with an asterisk (*) are required.
        - If a field is not applicable, leave it blank.

        **Experiment ID Formatting:**
        - `ExperimentType_Index`  &nbsp; (e.g., `HPHT_001`, `SERUM_020`)
        - Longer strings before the underscore (e.g., `Serum_MH_020`)
        - Sequential runs are indicated by adding a number after the underscore (e.g., `Serum_MH_101-2`)
        - Treatment variants are indicated by adding a text after the underscore (e.g., `Serum_MH_101_Desorption`)
        - Combined experiments are indicated by adding a number and a text after the underscore (e.g., `Serum_MH_101-2_Desorption`)

        Please fill out each sheet as appropriate for your experiments. For details on additional naming options or behaviors (sequential runs, renaming, copying from parent, etc.), see the detailed documentation.
        """
    )

    # Build template
    unit_options = [u.value for u in AmountUnit]
    experiments_cols = [
        "experiment_id",  # required
        "old_experiment_id",  # optional: use when renaming experiments
        "sample_id",
        "date",           # any Excel/ISO date
        "status",         # ExperimentStatus name or value
        "initial_note",   # first ExperimentNotes entry
        "overwrite",      # True/False
    ]

    # Multiple example rows showing different naming patterns (both 2-part and 3-part formats)
    example_experiments = [
        {
            "experiment_id": "Serum_MH_101",
            "sample_id": "Rock_1",
            "date": pd.Timestamp.today().date(),
            "status": "ONGOING",
            "initial_note": "Base experiment (3-part format)",
            "overwrite": False,
        },
        {
            "experiment_id": "HPHT_001",
            "sample_id": "Rock_1",
            "date": pd.Timestamp.today().date(),
            "status": "ONGOING",
            "initial_note": "Base experiment (2-part format)",
            "overwrite": False,
        },
        {
            "experiment_id": "Serum_MH_101-2",
            "sample_id": "Rock_1",
            "date": pd.Timestamp.today().date(),
            "status": "ONGOING",
            "initial_note": "2nd consecutive run (sequential, 3-part)",
            "overwrite": False,
        },
        {
            "experiment_id": "HPHT_001-2",
            "sample_id": "Rock_1",
            "date": pd.Timestamp.today().date(),
            "status": "ONGOING",
            "initial_note": "2nd consecutive run (sequential, 2-part)",
            "overwrite": False,
        },
        {
            "experiment_id": "Serum_MH_101_Desorption",
            "sample_id": "Rock_1",
            "date": pd.Timestamp.today().date(),
            "status": "ONGOING",
            "initial_note": "Treatment variant (3-part)",
            "overwrite": False,
        },
        {
            "experiment_id": "HPHT_001_Desorption",
            "sample_id": "Rock_1",
            "date": pd.Timestamp.today().date(),
            "status": "ONGOING",
            "initial_note": "Treatment variant (2-part)",
            "overwrite": False,
        },
        {
            "experiment_id": "Serum_MH_101-2_Annealing",
            "sample_id": "Rock_1",
            "date": pd.Timestamp.today().date(),
            "status": "ONGOING",
            "initial_note": "Combined: 2nd run + treatment (3-part)",
            "overwrite": False,
        },
        {
            "experiment_id": "HPHT_001-2_Annealing",
            "sample_id": "Rock_1",
            "date": pd.Timestamp.today().date(),
            "status": "ONGOING",
            "initial_note": "Combined: 2nd run + treatment (2-part)",
            "overwrite": False,
        },
    ]

    # Derive conditions headers from model columns to avoid hardcoding
    # We exclude PK/FKs/metadata
    from database import ExperimentalConditions as _EC
    cond_reserved = {"id", "experiment_id", "experiment_fk", "created_at", "updated_at"}
    cond_blacklist = {
        "catalyst", "catalyst_mass",
        "buffer_system", "buffer_concentration",
        "surfactant_type", "surfactant_concentration",
        "catalyst_percentage", "catalyst_ppm",
        "water_to_rock_ratio",  # Calculated field
    }
    conditions_cols = [
        c.name for c in _EC.__table__.columns
        if c.name not in cond_reserved and c.name not in cond_blacklist
    ]

    additives_cols = ["experiment_id", "compound", "amount", "unit", "order", "method"]
    example_add = {
        "experiment_id": example_experiments[0]["experiment_id"],
        "compound": "Sodium Chloride",
        "amount": 100.0,
        "unit": unit_options[0] if unit_options else "mg",
        "order": 1,
        "method": "solution",
    }

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # INSTRUCTIONS sheet
        instructions_data = {
            "Naming Convention": [
                "Base Experiment (3-part)",
                "Base Experiment (2-part)",
                "Sequential Run (3-part)",
                "Sequential Run (2-part)",
                "Treatment Variant (3-part)",
                "Treatment Variant (2-part)",
                "Combined (3-part)",
                "Combined (2-part)",
                "",
                "IMPORTANT NOTES",
                "",
                "AUTO-COPY FEATURE",
                "",
                "Override Behavior",
            ],
            "Pattern": [
                "TYPE_INITIALS_INDEX",
                "TYPE_INDEX",
                "TYPE_INITIALS_INDEX-NUMBER",
                "TYPE_INDEX-NUMBER",
                "TYPE_INITIALS_INDEX_TREATMENT",
                "TYPE_INDEX_TREATMENT",
                "TYPE_INITIALS_INDEX-NUMBER_TREATMENT",
                "TYPE_INDEX-NUMBER_TREATMENT",
                "",
                "",
                "",
                "",
                "",
                "",
            ],
            "Example": [
                "Serum_MH_101",
                "HPHT_001",
                "Serum_MH_101-2",
                "HPHT_001-2",
                "Serum_MH_101_Desorption",
                "HPHT_001_Desorption",
                "Serum_MH_101-2_Desorption",
                "HPHT_001-2_Desorption",
                "",
                "",
                "",
                "",
                "",
                "",
            ],
            "Description": [
                "Initial/base experiment (3-part format)",
                "Initial/base experiment (2-part format)",
                "2nd, 3rd... consecutive run (use hyphen-NUMBER)",
                "2nd, 3rd... consecutive run (use hyphen-NUMBER)",
                "Special treatment on sample (use underscore_TEXT)",
                "Special treatment on sample (use underscore_TEXT)",
                "Treatment on specific run's sample",
                "Treatment on specific run's sample",
                "",
                "Hyphens track numeric lineage. Underscores indicate treatments.",
                "Both 2-part (TYPE_INDEX) and 3-part (TYPE_INITIALS_INDEX) formats supported.",
                "Sequential/treatment: CONDITIONS auto-copy from parent (overwrite=False)",
                "Additives NEVER auto-copy - must be explicitly provided each time",
                "Researcher (3-part only), sample_id, experiment_type extracted from experiment_id",
            ],
        }
        df_instructions = pd.DataFrame(instructions_data)
        df_instructions.to_excel(writer, index=False, sheet_name='INSTRUCTIONS_READ_FIRST')
        
        # experiments sheet with multiple examples
        df_exp = pd.DataFrame(example_experiments, columns=experiments_cols)
        # Display: mark required fields with asterisks and add format hint
        display_experiments_cols = [
            "experiment_id* (TYPE_INDEX or TYPE_INITIALS_INDEX)", 
            "sample_id", 
            "date", 
            "status", 
            "initial_note", 
            "old_experiment_id (optional, for renames)",
            "overwrite"
        ]
        df_exp.columns = display_experiments_cols
        df_exp.to_excel(writer, index=False, sheet_name='experiments')
        
        # conditions: one example row with blanks
        df_cond = pd.DataFrame([{c: None for c in ["experiment_id"] + conditions_cols}]).assign(experiment_id=example_experiments[0]["experiment_id"]) 
        # Display: mark required experiment_id with asterisk
        display_conditions_cols = ["experiment_id*"] + conditions_cols
        df_cond.columns = display_conditions_cols
        df_cond.to_excel(writer, index=False, sheet_name='conditions')
        
        # additives sheet
        df_add = pd.DataFrame([example_add], columns=additives_cols)
        # Display: mark required columns with asterisks
        display_additives_cols = ["experiment_id*", "compound*", "amount*", "unit*", "order", "method"]
        df_add.columns = display_additives_cols
        df_add.to_excel(writer, index=False, sheet_name='additives')

        # Autosize columns for readability
        try:
            from frontend.components.utils import autosize_excel_columns
            autosize_excel_columns(writer, 'INSTRUCTIONS_READ_FIRST')
            autosize_excel_columns(writer, 'experiments')
            autosize_excel_columns(writer, 'conditions')
            autosize_excel_columns(writer, 'additives')
        except Exception:
            pass
    buf.seek(0)

    st.download_button(
        label="Download New Experiments Template",
        data=buf,
        file_name="new_experiments_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    uploaded = st.file_uploader("Upload filled New Experiments template (xlsx)", type=["xlsx"])

    if not uploaded:
        return

    db = SessionLocal()
    try:
        created, updated, skipped, errors, warnings, info_messages = NewExperimentsUploadService.bulk_upsert_from_excel(db, uploaded.read())
        if errors:
            db.rollback()
            st.error("Upload encountered issues; no changes were applied.")
            for msg in errors[:50]:
                st.error(msg)
            if len(errors) > 50:
                st.info(f"...and {len(errors)-50} more errors")
        else:
            db.commit()
            st.success(f"Experiments created: {created}, updated: {updated}, skipped rows: {skipped}")
            
            # Display info messages (including DEBUG messages, renames, and auto-copy info)
            if info_messages:
                # Count different message types
                debug_count = len([msg for msg in info_messages if 'DEBUG' in msg])
                rename_count = len([msg for msg in info_messages if 'Renamed experiment' in msg or 'Will rename' in msg])
                copy_count = len([msg for msg in info_messages if 'Will copy from parent' in msg or 'Copied' in msg])
                
                # Build expander title
                title_parts = []
                if debug_count > 0:
                    title_parts.append(f"{debug_count} DEBUG messages")
                if rename_count > 0:
                    title_parts.append(f"{rename_count} renames")
                if copy_count > 0:
                    title_parts.append(f"{copy_count} auto-copies")
                
                title = f"ℹ️ {' | '.join(title_parts) if title_parts else f'{len(info_messages)} info messages'} - click to view details"
                
                with st.expander(title, expanded=True):
                    for msg in info_messages[:200]:
                        st.info(msg)
                    if len(info_messages) > 200:
                        st.info(f"...and {len(info_messages)-200} more messages")
            
            # Display warnings (non-blocking)
            if warnings:
                with st.expander(f"⚠️ {len(warnings)} validation warning(s) - click to view", expanded=False):
                    for msg in warnings[:100]:
                        st.warning(msg)
                    if len(warnings) > 100:
                        st.info(f"...and {len(warnings)-100} more warnings")
    except Exception as e:
        db.rollback()
        st.error(f"Unexpected error during New Experiments upload: {e}")
    finally:
        db.close()


def handle_pxrf_upload():
    """Bulk upload pXRF Excel via backend service. No template required."""
    st.header("Bulk Upload pXRF Readings (Excel)")
    st.markdown(
        """
        Upload the pXRF Excel exported from the instrument/lab. The importer expects:
        - A `Reading No` column and element concentration columns per row
        - Optional: set "Update existing" to overwrite matching readings
        """
    )

    update_existing = st.checkbox("Update existing readings if present", value=False)
    uploaded = st.file_uploader("Upload pXRF Excel", type=["xlsx", "xlsm", "xls"])
    if not uploaded:
        return

    db = SessionLocal()
    try:
        inserted, updated, skipped, errors = PXRFUploadService.ingest_from_bytes(db, uploaded.read(), update_existing=update_existing)
        if errors:
            db.rollback()
            st.error("Upload encountered issues; no changes were applied.")
            for msg in errors[:50]:
                st.error(msg)
            if len(errors) > 50:
                st.info(f"...and {len(errors)-50} more errors")
            return
        db.commit()
        st.success(f"pXRF readings — inserted: {inserted}, updated: {updated}, skipped: {skipped}")
    except Exception as e:
        db.rollback()
        st.error(f"Unexpected error during pXRF upload: {e}")
    finally:
        db.close()

def handle_rock_samples_upload():
    """Bulk upload rock inventory and attach photos."""
    st.header("Bulk Upload Rock Inventory")
    st.markdown(
        """
        Use the template to add or update rock samples. Columns include `sample_id` (required),
        classification and location fields. Optionally upload photos whose filenames match `sample_id`.
        
        **Overwrite Mode:** Set `overwrite` to `True` for a sample to completely replace all existing fields.
        When `False` or blank, only provided fields are updated (existing fields remain unchanged).
        """
    )

    # Template
    template_df = pd.DataFrame([
        {
            "sample_id": "Rock_1",
            "rock_classification": "Basalt",
            "state": "CO",
            "country": "USA",
            "locality": "Somewhere",
            "latitude": 39.7392,
            "longitude": -104.9903,
            "description": "Dark fine-grained rock",
            "pxrf_reading_no": "12345A",
            "overwrite": False,
        }
    ], columns=[
        "sample_id", "rock_classification", "state", "country", "locality",
        "latitude", "longitude", "description", "pxrf_reading_no", "overwrite"
    ])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        template_df.to_excel(writer, index=False, sheet_name='samples')
        
        # Autosize columns for readability
        try:
            from frontend.components.utils import autosize_excel_columns
            autosize_excel_columns(writer, 'samples')
        except Exception:
            pass
    buf.seek(0)
    st.download_button(
        label="Download Rock Inventory Template",
        data=buf,
        file_name="rock_inventory_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    
    # Add overwrite checkbox
    overwrite_all = st.checkbox(
        "Overwrite all existing sample fields",
        value=False,
        help="When checked, replaces ALL fields for existing samples. Otherwise, only updates provided fields. Per-row 'overwrite' column takes precedence."
    )
    
    uploaded = st.file_uploader("Upload filled rock inventory template (xlsx)", type=["xlsx"])
    images = st.file_uploader("Optional: Upload sample photos (filenames must match sample_id)", type=["jpg", "jpeg", "png"], accept_multiple_files=True)

    if not uploaded:
        return

    # Prepare images list
    image_tuples = []
    if images:
        for img in images:
            try:
                image_tuples.append((img.name, img.read(), img.type))
            except Exception:
                continue

    db = SessionLocal()
    try:
        created, updated, images_attached, skipped, errors, warnings = RockInventoryService.bulk_upsert_samples(
            db, uploaded.read(), image_tuples, overwrite_all=overwrite_all
        )
        if errors:
            db.rollback()
            st.error("Upload encountered issues; no changes were applied.")
            for msg in errors[:50]:
                st.error(msg)
            if len(errors) > 50:
                st.info(f"...and {len(errors)-50} more errors")
        else:
            db.commit()
            st.success(f"Rock inventory — created: {created}, updated: {updated}, images attached: {images_attached}, skipped rows: {skipped}")
            
            # Display warnings if any (non-blocking)
            if warnings:
                with st.expander(f"⚠️ {len(warnings)} warning(s) - click to view", expanded=False):
                    for msg in warnings[:100]:
                        st.warning(msg)
                    if len(warnings) > 100:
                        st.info(f"...and {len(warnings)-100} more warnings")
    except Exception as e:
        db.rollback()
        st.error(f"Unexpected error during rock inventory upload: {e}")
    finally:
        db.close()

def handle_actlabs_titration_upload():
    """Thin UI wrapper to import ActLabs Excel via backend service."""
    st.header("Bulk Upload ActLabs Rock Titration Sheet")
    st.markdown(
        """
        Upload the raw ActLabs report (Excel or CSV). The importer auto-detects headers and formats
        and loads titration results per sample.
        """
    )

    uploaded = st.file_uploader("Upload ActLabs report (Excel or CSV)", type=["xlsx", "xlsm", "xls", "csv"])
    if not uploaded:
        return

    # Read file once and reuse for diagnostics and import
    file_bytes = uploaded.read()

    # Diagnostics preview
    with st.spinner("Analyzing file structure..."):
        diags, warns = ActlabsRockTitrationService.diagnose(file_bytes)

    if warns:
        st.subheader("⚠️ Detected Issues / Warnings")
        for w in warns:
            st.warning(w)

    if diags:
        st.subheader("🧭 Detected Layout")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.metric("Read mode", str(diags.get("read_mode", "?")))
        with col2:
            shp = diags.get("shape") or (0, 0)
            st.metric("Shape (rows, cols)", f"{shp[0]}, {shp[1]}")
        with col3:
            st.metric("Data start row (0-based)", str(diags.get("data_start_row", "?")))

        st.caption(f"Sample ID column index: {diags.get('sample_id_col', '?')}")

        analytes = diags.get("analytes", [])
        if analytes:
            st.write("Analytes detected (last header occurrence wins):")
            st.dataframe(pd.DataFrame(analytes))
        else:
            st.info("No analyte headers detected.")

        sample_preview = diags.get("sample_id_preview", [])
        if sample_preview:
            st.write("Sample ID preview:")
            st.dataframe(pd.DataFrame({"sample_id": sample_preview}))
        else:
            st.info("No sample IDs detected in data preview.")

        value_quality = diags.get("analyte_value_quality", [])
        if value_quality:
            st.write("Value quality check (first rows):")
            st.dataframe(pd.DataFrame(value_quality))

    can_import = bool(diags) and bool(diags.get("analytes")) and bool(diags.get("sample_id_preview"))
    if not can_import:
        st.info("Import disabled until at least one analyte and one sample ID are detected.")

    if st.button("Import Parsed Data", disabled=not can_import):
        db = SessionLocal()
        try:
            created, updated, skipped, errors = ActlabsRockTitrationService.import_excel(db, file_bytes)
            if errors:
                db.rollback()
                st.error("Upload encountered issues; no changes were applied.")
                for msg in errors[:50]:
                    st.error(msg)
                if len(errors) > 50:
                    st.info(f"...and {len(errors)-50} more errors")
            else:
                db.commit()
                st.success(f"Imported ActLabs titration results — created: {created}, updated: {updated}, skipped rows: {skipped}")
        except Exception as e:
            db.rollback()
            st.error(f"Unexpected error during ACTLABS upload: {e}")
        finally:
            db.close()

## analyte upload removed: analytes are created during ACTLABS titration upload


def handle_elemental_composition_upload():
    """Thin UI wrapper that delegates composition upsert to backend service."""
    st.header("Bulk Upload Elemental Composition (Titration/ICP)")
    st.markdown(
        """
        Download the template to enter elemental compositions per sample. The sheet contains `sample_id`
        and one column per analyte. Upload the filled template to upsert composition values.
        """
    )

    # Build template based on existing analytes via backend-friendly DB call
    db = SessionLocal()
    try:
        # Lazy import to avoid circulars; keep simple here
        from database import Analyte as _Analyte
        analytes = db.query(_Analyte).order_by(_Analyte.analyte_symbol.asc()).all()
        analyte_cols = [a.analyte_symbol for a in analytes] if analytes else ["FeO", "SiO2"]
    finally:
        db.close()

    cols = ["sample_id"] + analyte_cols
    template_df = pd.DataFrame([{c: None for c in cols}])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        template_df.to_excel(writer, index=False, sheet_name='composition')
        
        # Autosize columns for readability
        try:
            from frontend.components.utils import autosize_excel_columns
            autosize_excel_columns(writer, 'composition')
        except Exception:
            pass
    buf.seek(0)
    st.download_button(
        label="Download Composition Template",
        data=buf,
        file_name="elemental_composition_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    uploaded = st.file_uploader("Upload filled composition template (xlsx)", type=["xlsx"])
    if not uploaded:
        return

    db = SessionLocal()
    try:
        created, updated, skipped, errors = ElementalCompositionService.bulk_upsert_wide_from_excel(db, uploaded.read())
        if errors:
            db.rollback()
            st.error("Upload failed; no changes were applied.")
            for msg in errors[:50]:
                st.error(msg)
            if len(errors) > 50:
                st.info(f"...and {len(errors)-50} more errors")
        else:
            db.commit()
            st.success(f"Compositions created: {created}, updated: {updated}, skipped rows: {skipped}")
    except Exception as e:
        db.rollback()
        st.error(f"Unexpected error during composition upload: {e}")
    finally:
        db.close()

def upload_chemical_inventory():
    """
    Upload a template-driven Excel to create or update chemical compounds in the database.
    - Provides a downloadable template with supported fields
    - Accepts an uploaded Excel and upserts rows into the `compounds` table
    """
    st.header("Bulk Upload Chemical Inventory (Compounds)")
    st.markdown(
        """
        Create or update compound records. `name` is required; other fields (formula, CAS, molecular weight,
        density, melting/boiling point, supplier, etc.) are optional. Later uploads can enrich existing entries.
        """
    )

    # Template definition (column order preserved)
    template_columns = [
        "name",                 # required, unique
        "formula",
        "cas_number",          # optional, unique if provided
        "molecular_weight",    # g/mol
        "density",             # g/cm³ (solids) or g/mL (liquids)
        "melting_point",       # °C
        "boiling_point",       # °C
        "solubility",
        "hazard_class",
        "supplier",
        "catalog_number",
        "notes",
    ]

    example_row = {
        "name": "Sodium Chloride",
        "formula": "NaCl",
        "cas_number": "7647-14-5",
        "molecular_weight": 58.44,
        "density": 2.165,
        "melting_point": 801.0,
        "boiling_point": 1413.0,
        "solubility": "Soluble in water",
        "hazard_class": "Non-hazardous",
        "supplier": "Sigma-Aldrich",
        "catalog_number": "S7653",
        "notes": "Food grade"
    }

    template_df = pd.DataFrame([example_row], columns=template_columns)

    # Download template
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        template_df.to_excel(writer, index=False, sheet_name='compounds')
        
        # Autosize columns for readability
        try:
            from frontend.components.utils import autosize_excel_columns
            autosize_excel_columns(writer, 'compounds')
        except Exception:
            pass
    buf.seek(0)
    st.download_button(
        label="Download Compound Template",
        data=buf,
        file_name="compound_inventory_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    uploaded_file = st.file_uploader("Upload filled compound template (xlsx)", type=["xlsx"])

    if not uploaded_file:
        return

    db = SessionLocal()
    try:
        created, updated, skipped, errors = ChemicalInventoryService.bulk_upsert_from_excel(db, uploaded_file.read())
        if errors:
            db.rollback()
            st.error("Upload failed; no changes were applied.")
            for msg in errors[:50]:
                st.error(msg)
            if len(errors) > 50:
                st.info(f"...and {len(errors)-50} more errors")
        else:
            db.commit()
            st.success(f"Compounds created: {created}, updated: {updated}, skipped: {skipped}")
    except Exception as e:
        db.rollback()
        st.error(f"Unexpected error during compound upload: {e}")
    finally:
        db.close()

def experiments_additives():
    """
    Bulk upload chemical additives for specific experiments.

    Template columns (sheet name 'experiment_additives'):
      - experiment_id* (string)
      - compound* (existing compound name)
      - amount* (numeric > 0)
      - unit* (one of AmountUnit values, e.g., g, mg, μg, kg, μL, mL, L, μmol, mmol, mol, ppm, mM, M, %, wt%)
      - order (optional integer)
      - method (optional text)
    """
    st.header("Bulk Upload Experiment Additives")
    st.markdown(
        """
        Add or update chemical additives for experiments. The sheet requires:
        - `experiment_id*`: existing experiment identifier
        - `compound*`: compound name (must exist in Chemical Inventory)
        - `amount*` and `unit*`: quantity and unit for the additive
        Optional: `order` and `method`.
        """
    )

    # Build template
    unit_options = [u.value for u in AmountUnit]
    example = {
        'experiment_id': 'Serum_MH_001',
        'compound': 'Sodium Chloride',
        'amount': 100.0,
        'unit': unit_options[0] if unit_options else 'mg',
        'order': 1,
        'method': 'solution'
    }
    template_df = pd.DataFrame([example], columns=['experiment_id', 'compound', 'amount', 'unit', 'order', 'method'])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        template_df.to_excel(writer, index=False, sheet_name='experiment_additives')
        
        # Autosize columns for readability
        try:
            from frontend.components.utils import autosize_excel_columns
            autosize_excel_columns(writer, 'experiment_additives')
        except Exception:
            pass
    buf.seek(0)
    st.download_button(
        label="Download Additives Template",
        data=buf,
        file_name="experiment_additives_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    uploaded = st.file_uploader("Upload filled additives template (xlsx)", type=["xlsx"])

    if not uploaded:
        return

    try:
        df = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Failed to read Excel: {e}")
        return

    db = SessionLocal()
    try:
        created, updated, skipped, errors = ExperimentAdditivesService.bulk_upsert_from_excel(db, uploaded.read())
        if errors:
            db.rollback()
            st.error("Upload failed; no changes were applied.")
            for msg in errors[:50]:
                st.error(msg)
            if len(errors) > 50:
                st.info(f"...and {len(errors)-50} more errors")
        else:
            db.commit()
            st.success(f"Additives created: {created}, updated: {updated}, skipped: {skipped}")
    except Exception as e:
        db.rollback()
        st.error(f"Unexpected error during additives upload: {e}")
    finally:
        db.close()

def handle_solution_chemistry_upload():
    """
    Handles the UI and logic for bulk uploading solution chemistry results (NMR-quantified).
    """
    st.header("Bulk Upload Solution Chemistry Results")
    st.markdown("""
    Upload an Excel file for results such as NMR, Hydrogen, pH, Conductivity, Alkalinity. 
    The file should have the following columns found in the template below.

    ***Instructions:***
    - Columns marked with an asterisk (*) are required.
    - If a field is not applicable, leave it blank.
    - Experiment ID, Description, and Measurement Date are required.
   
    ***Download the template below to get started.***
    """)

    # --- Template Generation ---
    # Use English headers from config
    template_data = {
        SCALAR_RESULTS_TEMPLATE_HEADERS["measurement_date"]: [pd.Timestamp.now().date()],
        SCALAR_RESULTS_TEMPLATE_HEADERS["experiment_id"]: ["Serum_MH_025"],
        SCALAR_RESULTS_TEMPLATE_HEADERS["time_post_reaction"]: [1],  # Optional field
        SCALAR_RESULTS_TEMPLATE_HEADERS["description"]: ["Sampled after acid addition"],
        SCALAR_RESULTS_TEMPLATE_HEADERS["gross_ammonium_concentration_mM"]: [10.5],
        SCALAR_RESULTS_TEMPLATE_HEADERS["sampling_volume_mL"]: [5.0],
        SCALAR_RESULTS_TEMPLATE_HEADERS["background_ammonium_concentration_mM"]: [0.0],
        SCALAR_RESULTS_TEMPLATE_HEADERS["background_experiment_id"]: [""],
        # Hydrogen gas sampling fields
        SCALAR_RESULTS_TEMPLATE_HEADERS["h2_concentration"]: [0.00],
        SCALAR_RESULTS_TEMPLATE_HEADERS["gas_sampling_volume_ml"]: [0.0],
        SCALAR_RESULTS_TEMPLATE_HEADERS["gas_sampling_pressure_MPa"]: [0.1013],
        SCALAR_RESULTS_TEMPLATE_HEADERS["final_ph"]: [7.2],
        SCALAR_RESULTS_TEMPLATE_HEADERS["ferrous_iron_yield"]: [0.0],
        SCALAR_RESULTS_TEMPLATE_HEADERS["final_nitrate_concentration_mM"]: [0],
        SCALAR_RESULTS_TEMPLATE_HEADERS["final_dissolved_oxygen_mg_L"]: [0],
        SCALAR_RESULTS_TEMPLATE_HEADERS["co2_partial_pressure_MPa"]: [0],
        SCALAR_RESULTS_TEMPLATE_HEADERS["final_conductivity_mS_cm"]: [1500.0],
        SCALAR_RESULTS_TEMPLATE_HEADERS["final_alkalinity_mg_L"]: [120.0],
        SCALAR_RESULTS_TEMPLATE_HEADERS["overwrite"]: [False],

    }

    # Helper to find original key from header value
    REVERSE_HEADERS = {v: k for k, v in SCALAR_RESULTS_TEMPLATE_HEADERS.items()}

    template_cols_ordered = {}
    for col, val in template_data.items():
        # Get the original variable name to check if it's required
        variable_name = REVERSE_HEADERS.get(col)
        if variable_name in EXPERIMENTAL_RESULTS_REQUIRED_COLS:
            template_cols_ordered[f"{col}*"] = val
        else:
            template_cols_ordered[col] = val
    
    template_df = pd.DataFrame(template_cols_ordered)

    # Create template Excel file using simple approach
    output = io.BytesIO()
    
    # Build instructions DataFrame
    instructions_data = {
        "Topic": [
            "Required Fields",
            "",
            "Overwrite Behavior",
            "",
            "",
            "",
            "",
            "",
            "Examples",
            "",
            "",
            "",
            "Important Notes",
            "",
            "",
        ],
        "Description": [
            "Columns marked with asterisk (*) in the data sheet are required",
            "",
            "Controls how existing data is updated when you upload results for experiments that already have data",
            "",
            "PARTIAL UPDATE (overwrite = False, default):",
            "  • Only columns you provide are updated",
            "  • Empty cells leave existing database values unchanged",
            "  • Best for adding new measurements or correcting specific values",
            "",
            "COMPLETE REPLACEMENT (overwrite = True):",
            "  • ALL fields are replaced with your uploaded values",
            "  • Empty cells will set database fields to blank/null",
            "  • Use when you want to completely replace an entire result entry",
            "",
            "Per-row 'overwrite' column takes precedence over global checkbox in UI",
        ],
        "Example/Value": [
            "experiment_id, description",
            "",
            "Use 'overwrite' column or checkbox",
            "",
            "False (or leave blank)",
            "",
            "",
            "",
            "",
            "True",
            "",
            "",
            "",
            "You can mix modes in one upload",
            "",
        ],
    }
    df_instructions = pd.DataFrame(instructions_data)
    
    # Write both sheets to Excel (avoid context manager in case that's the issue)
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    try:
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Add INSTRUCTIONS sheet
        ws1 = wb.create_sheet('INSTRUCTIONS_READ_FIRST')
        for r_idx, row in enumerate(dataframe_to_rows(df_instructions, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws1.cell(row=r_idx, column=c_idx, value=value)
        
        # Add Solution Chemistry sheet
        ws2 = wb.create_sheet('Solution Chemistry')
        for r_idx, row in enumerate(dataframe_to_rows(template_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=value)
        
        # Ensure sheets are visible
        ws1.sheet_state = 'visible'
        ws2.sheet_state = 'visible'
        
        # Save to BytesIO
        wb.save(output)
        output.seek(0)
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error creating solution chemistry Excel template: {e}", exc_info=True)
        st.error(f"Error creating template: {e}")
        # Return a minimal valid Excel file to prevent further errors
        output = io.BytesIO()
        wb_fallback = Workbook()
        ws_fallback = wb_fallback.active
        ws_fallback.title = "Error"
        ws_fallback['A1'] = "Template generation failed"
        wb_fallback.save(output)
        output.seek(0)

    st.download_button(
        label="Download Template",
        data=output,
        file_name="solution_chemistry_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.markdown("---")
    
    # Add overwrite checkbox
    overwrite_all = st.checkbox(
        "Overwrite all existing result fields",
        value=False,
        help="When checked, replaces ALL fields for existing results. Otherwise, only updates provided fields. Per-row 'overwrite' column takes precedence."
    )
    
    uploaded_file = st.file_uploader("Upload your Excel file", type=["xlsx"])

    if uploaded_file:
        db = SessionLocal()
        try:
            created, updated, skipped, errors = ScalarResultsUploadService.bulk_upsert_from_excel(db, uploaded_file.read(), overwrite_all=overwrite_all)
            if errors:
                db.rollback()
                for error in errors:
                    st.warning(error)
                st.error("Upload failed due to errors. Please correct the file and try again.")
            else:
                db.commit()
                st.success(f"Successfully uploaded {created} solution chemistry results.")
        except IntegrityError as e:
            db.rollback()
            st.error(f"Database error: {e.orig}")
        except Exception as e:
            db.rollback()
            st.error(f"An unexpected error occurred: {e}")
        finally:
            db.close()

def _process_solution_chemistry_df(df):
    # Deprecated: logic moved to backend/services/bulk_uploads/scalar_results.py
    pass

def handle_icp_upload():
    """
    Handles the UI and logic for bulk uploading ICP-OES elemental analysis results.
    Processes CSV files directly from ICP-OES instrument output.
    """
    st.header("Bulk Upload ICP-OES Elemental Analysis")
    st.markdown("""
    Upload a CSV file containing ICP-OES elemental analysis data. 
    This should be the direct output from the ICP-OES instrument containing:
    
    - **Experiment ID**, **Time Point**, and **Dilution** in 'Label' Column (e.g. 'Serum_MH_011_Day5_5x)
    - **Elemental concentrations** in ppm (Fe, Mg, Ni, Cu, Si, Co, Mo, Al, etc.)
    - Please ensure all referenced Experiment IDs exist in the database
    
    **Note:** No template is provided as this processes direct instrument output files. This is for ICP-OES data only.
    """)

    uploaded_file = st.file_uploader("Upload your CSV file", type=["csv"])

    if uploaded_file:
        # Show CSV analysis and allow manual header row specification
        file_content = uploaded_file.read()
        uploaded_file.seek(0)  # Reset for potential re-reading
        
        # Quick analysis
        diagnosis = ICPService.diagnose_csv_structure(file_content)
        
        if 'error' not in diagnosis:
            st.subheader("📋 CSV File Analysis")
            st.write(f"**Total lines:** {diagnosis['total_lines']}")
            
            # Show header candidates
            header_candidates = [line for line in diagnosis['line_analysis'] 
                               if line['has_icp_keywords'] and 5 <= line['column_count'] <= 50]
            
            if header_candidates:
                st.write("**Detected header candidates:**")
                for candidate in header_candidates[:5]:  # Show top 5 candidates
                    st.write(f"Line {candidate['line_number']}: {candidate['column_count']} columns")
                    st.caption(f"Preview: {candidate['preview']}")
            
            # Manual override option
            if len(header_candidates) > 1:
                st.warning("Multiple header candidates detected. You can manually specify the header row if needed.")
                manual_header = st.number_input(
                    "Manual header row (0-based, leave 0 for auto-detection):", 
                    min_value=0, 
                    max_value=min(20, diagnosis['total_lines']-1), 
                    value=0
                )
            else:
                manual_header = 0
        else:
            manual_header = 0
        
        if st.button("Process ICP Data"):
            try:
                _process_icp_csv(file_content, manual_header)
                
            except Exception as e:
                st.error(f"An error occurred processing the CSV file: {e}")

def _process_icp_csv(file_content: bytes, manual_header_row: int = 0):
    """
    Processes the ICP-OES CSV file using ICPService for all backend logic.
    """
    db = SessionLocal()
    try:
        # Step 1: Parse and process the ICP file
        st.info("Processing ICP-OES data file...")
        processed_data, processing_errors = ICPService.parse_and_process_icp_file(file_content, manual_header_row)
        
        if processing_errors:
            st.subheader("⚠️ Processing Issues Found")
            for error in processing_errors:
                st.warning(error)
        
        if not processed_data:
            st.error("No valid ICP-OES data found to upload.")
            return
        
        # Step 2: Show preview of processed data
        st.subheader("📊 Processed ICP-OES Data Preview")
        
        # Convert processed data to DataFrame for display
        preview_df = pd.DataFrame(processed_data)
        
        # Show summary info
        st.info(f"Found {len(processed_data)} samples with ICP-OES data")
        
        # Show first few samples
        if len(preview_df) > 0:
            st.write("**Sample Overview:**")
            display_cols = ['experiment_id', 'time_post_reaction', 'dilution_factor']
            element_cols = [col for col in preview_df.columns if col not in ['experiment_id', 'time_post_reaction', 'dilution_factor', 'raw_label']]
            display_cols.extend(element_cols[:5])  # Show first 5 elements
            
            if len(element_cols) > 5:
                st.write(f"*Showing first 5 elements. Total elements detected: {len(element_cols)}*")
            
            st.dataframe(preview_df[display_cols].head(10))
        
        # Step 3: Upload to database if no critical errors
        if not processing_errors or st.button("Upload Despite Warnings"):
            st.info("Uploading ICP-OES data to database...")
            
            results, upload_errors = ICPService.bulk_create_icp_results(db, processed_data)
            
            # Separate errors from overwrite notifications
            actual_errors = [msg for msg in upload_errors if not msg.startswith("Sample") or "Updated existing ICP-OES data" not in msg]
            overwrite_notifications = [msg for msg in upload_errors if "Updated existing ICP-OES data" in msg]
            
            # Handle upload feedback
            if actual_errors:
                st.subheader("❌ Upload Errors")
                for error in actual_errors:
                    st.error(error)
            
            if overwrite_notifications:
                st.subheader("🔄 Data Updates")
                for notification in overwrite_notifications:
                    st.info(notification.replace("Sample ", "Entry "))
            
            if results and not actual_errors:
                db.commit()
                st.success(f"✅ Successfully processed {len(results)} ICP-OES results.")
                
                # Show success summary
                experiments = list(set([data['experiment_id'] for data in processed_data if data['experiment_id']]))
                new_count = len(results) - len(overwrite_notifications)
                update_count = len(overwrite_notifications)
                
                summary_parts = []
                if new_count > 0:
                    summary_parts.append(f"{new_count} new")
                if update_count > 0:
                    summary_parts.append(f"{update_count} updated")
                
                summary = " and ".join(summary_parts)
                st.info(f"**Summary:** {summary} ICP-OES results for {len(experiments)} experiments")
                
            elif not results and not actual_errors:
                st.info("ℹ️ No new ICP-OES data to upload.")
            else:
                st.error("❌ Upload failed due to errors. Please correct the issues and try again.")
                db.rollback()
        
    except IntegrityError as e:
        db.rollback()
        st.error(f"Database integrity error: {e.orig}")
    except Exception as e:
        db.rollback()
        st.error(f"An unexpected error occurred during ICP-OES processing: {e}")
    finally:
        db.close()

def handle_experiment_status_update():
    """
    Bulk update experiment status: mark listed experiments as ONGOING,
    mark all other ONGOING experiments as COMPLETED.
    Shows preview before applying changes with confirmation required.
    """
    st.header("Bulk Update Experiment Status")
    st.markdown("""
    Upload an Excel file with experiment IDs to mark as **ONGOING**. 
    All other **HPHT** experiments currently marked as **ONGOING** will be changed to **COMPLETED**.
    
    **Columns:**
    - `experiment_id` (required): Experiments to mark as ONGOING
    - `reactor_number` (optional): Update reactor number for HPHT experiments
    
    **Important Notes:**
    - Only HPHT experiments not in the list will be auto-completed
    - Other experiment types (Serum, Autoclave, etc.) maintain their status
    - You will see a preview before any changes are applied
    - Confirmation is required before applying changes
    """)

    # Generate template
    template_df = pd.DataFrame([
        {"experiment_id": "HPHT_MH_001", "reactor_number": 1},
        {"experiment_id": "HPHT_MH_002", "reactor_number": 2},
        {"experiment_id": "Serum_JD_015", "reactor_number": ""},
    ], columns=["experiment_id", "reactor_number"])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        template_df.to_excel(writer, index=False, sheet_name='experiment_status')
        
        # Autosize columns for readability
        try:
            from frontend.components.utils import autosize_excel_columns
            autosize_excel_columns(writer, 'experiment_status')
        except Exception:
            pass
    buf.seek(0)

    st.download_button(
        label="Download Experiment Status Template",
        data=buf,
        file_name="experiment_status_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.markdown("---")
    
    # File uploader
    uploaded_file = st.file_uploader("Upload filled template (xlsx)", type=["xlsx"], key="status_upload")

    if not uploaded_file:
        return

    # Get preview of changes
    db = SessionLocal()
    try:
        preview = ExperimentStatusService.preview_status_changes_from_excel(db, uploaded_file.read())
        
        # Show errors if any
        if preview.errors:
            st.error("Validation errors found:")
            for error in preview.errors:
                st.error(error)
            return
        
        # Show missing IDs warning
        if preview.missing_ids:
            st.warning(f"⚠️ {len(preview.missing_ids)} experiment ID(s) not found in database:")
            missing_df = pd.DataFrame({"Missing Experiment IDs": preview.missing_ids})
            st.dataframe(missing_df, use_container_width=True)
            st.info("These IDs will be skipped. Only experiments found in the database will be processed.")
        
        # Show preview of changes
        st.subheader("📋 Preview of Changes")
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("Experiments → ONGOING", len(preview.to_ongoing))
            if preview.to_ongoing:
                st.write("**Will be marked as ONGOING:**")
                ongoing_df = pd.DataFrame(preview.to_ongoing)
                # Reorder columns for better display
                col_order = ['experiment_id', 'current_status', 'current_reactor_number', 'new_reactor_number']
                ongoing_df = ongoing_df[[col for col in col_order if col in ongoing_df.columns]]
                st.dataframe(ongoing_df, use_container_width=True)
        
        with col2:
            st.metric("Experiments → COMPLETED", len(preview.to_completed))
            if preview.to_completed:
                st.write("**Will be marked as COMPLETED (HPHT only):**")
                completed_df = pd.DataFrame(preview.to_completed)
                # Reorder columns for better display
                col_order = ['experiment_id', 'current_status', 'current_reactor_number']
                completed_df = completed_df[[col for col in col_order if col in completed_df.columns]]
                st.dataframe(completed_df, use_container_width=True)
        
        # No changes to apply
        if not preview.to_ongoing and not preview.to_completed:
            st.info("ℹ️ No status changes needed based on this file.")
            return
        
        # Confirmation section
        st.markdown("---")
        st.subheader("⚠️ Confirm Changes")
        
        total_changes = len(preview.to_ongoing) + len(preview.to_completed)
        st.warning(f"This will update the status of **{total_changes} experiment(s)**. This action cannot be undone through the UI.")
        
        # Confirmation text input
        confirmation_text = st.text_input(
            'Type "CONFIRM" to proceed with status updates:',
            key="status_confirm"
        )
        
        # Apply button (disabled unless confirmation entered)
        apply_disabled = confirmation_text.upper() != "CONFIRM"
        
        if st.button("Apply Status Changes", disabled=apply_disabled, type="primary"):
            # Extract experiment IDs to mark as ONGOING
            exp_ids_to_ongoing = [exp["experiment_id"] for exp in preview.to_ongoing]
            
            # Get reactor_number_map from preview
            reactor_number_map = getattr(preview, 'reactor_number_map', {})
            
            # Apply changes
            marked_ongoing, marked_completed, reactor_updates, errors = ExperimentStatusService.apply_status_changes(
                db, exp_ids_to_ongoing, reactor_number_map
            )
            
            if errors:
                db.rollback()
                st.error("Failed to apply status changes:")
                for error in errors:
                    st.error(error)
            else:
                db.commit()
                st.success(f"✅ Status changes applied successfully!")
                st.info(f"**{marked_ongoing}** experiment(s) marked as ONGOING")
                st.info(f"**{marked_completed}** experiment(s) marked as COMPLETED")
                if reactor_updates > 0:
                    st.info(f"**{reactor_updates}** reactor number(s) updated")
                
                # Clear the confirmation text
                if 'status_confirm' in st.session_state:
                    del st.session_state['status_confirm']
        
        if apply_disabled and confirmation_text:
            st.info('Please type "CONFIRM" exactly (case-insensitive) to enable the Apply button.')
            
    except Exception as e:
        db.rollback()
        st.error(f"An unexpected error occurred: {e}")
    finally:
        db.close()